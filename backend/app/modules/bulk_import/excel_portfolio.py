"""Read a multi-sheet company/portfolio Excel workbook into the same in-memory
shape the JSON upload path already uses (see worker._process_json_portfolio
and mapping.py), with one addition: each portfolio dict also carries a
'paragraphs' list (the source document's paragraph-by-paragraph text, in
original order) built from the workbook's raw-body sheet.

This lets the rest of the pipeline -- company creation, complex matching,
image download, space grouping -- stay identical regardless of whether the
admin uploaded JSON or this Excel format. Only the "parse the uploaded file
into portfolio dicts" step differs.

Why this file needs to exist at all: the JSON export currently in use only
carries text already merged per detected space (`spaces[].description`),
because that merge happens upstream in the scraping/export tool. The Excel
export from the same tool keeps the pre-merge, paragraph-by-paragraph rows
(one row per paragraph, with its own document_order and sub_space_name) --
that granularity is what confidence.py needs to check whether a paragraph
ended up positioned where its matching photos actually are. Until the JSON
export is updated to carry that same granularity, Excel is the only format
that has it.
"""

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ExcelPortfolioFormatError(ValueError):
    pass


def _normalized(value: Any) -> str:
    return re.sub(r"[\s_\-·0-9]+", "", str(value or "")).strip()


def _str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def _float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# Each sheet type is identified by keyword(s) that must appear in its tab
# name. If tab names don't match any of these (e.g. a generic "Sheet1"-style
# export), fall back to matching by the sheet's own header row instead.
_SHEET_KEYWORDS: dict[str, tuple[str, ...]] = {
    "company": ("업체", "company"),
    "portfolio": ("포트폴리오", "portfolio"),
    "space": ("공간", "space"),
    "image": ("이미지", "image"),
    "paragraph": ("본문", "본문원문", "paragraph", "body"),
    "tag": ("태그", "tag"),
}

# Used only as a fallback when tab names don't carry a recognizable keyword
# (older exports used generic "Sheet1".."Sheet6" names). Matched against the
# sheet's own header row so the reader still works either way.
_SHEET_HEADER_SIGNATURES: dict[str, set[str]] = {
    "company": {"expertid", "companyname", "phonestatus"},
    "portfolio": {"portfolioid", "sourceurl", "areapyeong"},
    "space": {"spaceorder", "subspacename", "afterimagecount"},
    "image": {"imageorder", "documentorder", "imageurl"},
    "paragraph": {"documentorder", "nodetype", "text"},
    "tag": {"category", "tag"},
}


def _identify_sheets(workbook: Any) -> dict[str, str]:
    """Map sheet type -> actual sheet name in this workbook."""
    resolved: dict[str, str] = {}
    for name in workbook.sheetnames:
        norm_name = _normalized(name).lower()
        for sheet_type, keywords in _SHEET_KEYWORDS.items():
            if sheet_type in resolved:
                continue
            if any(kw in norm_name for kw in keywords):
                resolved[sheet_type] = name
                break

    missing = [t for t in _SHEET_KEYWORDS if t not in resolved]
    if missing:
        for name in workbook.sheetnames:
            if name in resolved.values():
                continue
            header = next(workbook[name].iter_rows(max_row=1, values_only=True), None)
            if not header:
                continue
            header_keys = {re.sub(r"[^a-z]", "", str(h or "").lower()) for h in header}
            for sheet_type in list(missing):
                signature = _SHEET_HEADER_SIGNATURES[sheet_type]
                if signature.issubset(header_keys):
                    resolved[sheet_type] = name
                    missing.remove(sheet_type)

    still_missing = [t for t in ("portfolio", "space", "image") if t not in resolved]
    if still_missing:
        raise ExcelPortfolioFormatError(
            "필수 시트를 찾을 수 없습니다 (portfolio/space/image): " + ", ".join(still_missing)
        )
    return resolved


def _read_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    keys = [str(h or "").strip() for h in headers]
    result = []
    for row in rows:
        if row is None or all(v is None for v in row):
            continue
        result.append({keys[i]: row[i] for i in range(min(len(keys), len(row)))})
    return result


def load_portfolio_workbook(path: Path) -> dict[str, Any]:
    """Parse the workbook into the same {'source', 'schema_version', 'companies',
    'portfolios'} shape the JSON path expects. Raises ExcelPortfolioFormatError
    on anything that isn't a readable, recognizable workbook."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelPortfolioFormatError("정상적인 Excel(.xlsx) 파일이 아닙니다.") from exc

    try:
        sheets = _identify_sheets(workbook)

        company_rows = _read_rows(workbook, sheets["company"]) if "company" in sheets else []
        portfolio_rows = _read_rows(workbook, sheets["portfolio"])
        space_rows = _read_rows(workbook, sheets["space"])
        image_rows = _read_rows(workbook, sheets["image"])
        paragraph_rows = _read_rows(workbook, sheets["paragraph"]) if "paragraph" in sheets else []
        tag_rows = _read_rows(workbook, sheets["tag"]) if "tag" in sheets else []
    finally:
        workbook.close()

    spaces_by_portfolio: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in space_rows:
        pid = _str(row.get("portfolio_id"))
        if not pid:
            continue
        spaces_by_portfolio[pid].append({
            "portfolio_id": pid,
            "space_order": _int(row.get("space_order")) or 0,
            "space_code": _str(row.get("space_code")) or "",
            "space_name": _str(row.get("space_name")) or "",
            "sub_space_code": _str(row.get("sub_space_code")) or "",
            "sub_space_name": _str(row.get("sub_space_name")) or "",
            "description": _str(row.get("description")) or "",
            "after_image_count": _int(row.get("after_image_count")) or 0,
            "before_image_count": _int(row.get("before_image_count")) or 0,
            "total_image_count": _int(row.get("total_image_count")) or 0,
        })

    images_by_portfolio: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in image_rows:
        pid = _str(row.get("portfolio_id"))
        if not pid:
            continue
        images_by_portfolio[pid].append({
            "portfolio_id": pid,
            "image_order": _int(row.get("image_order")),
            "document_order": _int(row.get("document_order")) or 0,
            "space_code": _str(row.get("space_code")) or "",
            "space_name": _str(row.get("space_name")) or "",
            "sub_space_code": _str(row.get("sub_space_code")) or "",
            "sub_space_name": _str(row.get("sub_space_name")) or "",
            "phase": _str(row.get("phase")) or "",
            "source_space_code": _str(row.get("source_space_code")),
            "classification_source": _str(row.get("classification_source")),
            "classification_confidence": _str(row.get("classification_confidence")),
            "image_url": _str(row.get("image_url")) or "",
            "width": _int(row.get("width")),
            "height": _int(row.get("height")),
            "ext_card_id": row.get("ext_card_id"),
            "keywords": row.get("keywords"),
            "alt": _str(row.get("alt")),
        })

    paragraphs_by_portfolio: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in paragraph_rows:
        pid = _str(row.get("portfolio_id"))
        if not pid:
            continue
        paragraphs_by_portfolio[pid].append({
            "document_order": _int(row.get("document_order")) or 0,
            "node_type": _str(row.get("node_type")) or "p",
            "space_code": _str(row.get("space_code")) or "",
            "sub_space_name": _str(row.get("sub_space_name")) or "",
            "text": _str(row.get("text")) or "",
        })
    for rows in paragraphs_by_portfolio.values():
        rows.sort(key=lambda r: r["document_order"])

    tags_by_portfolio: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in tag_rows:
        pid = _str(row.get("portfolio_id"))
        if not pid:
            continue
        tags_by_portfolio[pid].append({
            "category": _str(row.get("category")),
            "tag": _str(row.get("tag")),
        })

    portfolios: list[dict[str, Any]] = []
    for row in portfolio_rows:
        pid = _str(row.get("portfolio_id"))
        if not pid:
            continue
        style_tags = [t["tag"] for t in tags_by_portfolio.get(pid, []) if t.get("category") == "STYLE"]
        construction_tags = [
            t["tag"] for t in tags_by_portfolio.get(pid, []) if t.get("category") == "CONSTRUCTION"
        ]
        portfolios.append({
            "portfolio_id": pid,
            "source": _str(row.get("source")),
            "source_url": _str(row.get("source_url")),
            "title": _str(row.get("title")),
            "writer_id": _str(row.get("writer_id")),
            "writer_nickname": _str(row.get("writer_nickname")),
            "writer_type": _str(row.get("writer_type")),
            "company_name": _str(row.get("company_name")),
            "expert_id": _str(row.get("expert_id")),
            "expert_url": _str(row.get("expert_url")),
            "residence_type": _str(row.get("residence_type")),
            "area_pyeong": _str(row.get("area_pyeong")),
            "area_type": _str(row.get("area_type")),
            "apartment_name": _str(row.get("apartment_name")),
            "street_address": _str(row.get("street_address")),
            # v5.3 크롤러부터 추가된 필드. street_address가 비어도(이번
            # 배치 기준 73%가 비어있음) 단지를 찾아볼 수 있는 검색용
            # 텍스트 -- JSON 업로드 경로(worker._process_json_portfolio)는
            # 원본 dict를 그대로 쓰기 때문에 이미 이 필드를 그대로 받고
            # 있었고 worker._ensure_portfolio_complex도 이미 이 필드를
            # 소비하도록 되어 있었는데(v2.5.1, 2026-08-21), Excel 경로는
            # 이 함수가 만드는 dict에 화이트리스트로 없어서 조용히
            # 빠지고 있었다 -- 두 업로드 경로 동등성이 깨져 있던 버그.
            "address_lookup_query": _str(row.get("address_lookup_query")),
            # v5.3에서 새로 추가된 신뢰도/출처/보조 필드 일괄 추가
            # (2026-08-22, 사용자가 전달한 필드 우선순위 규칙 문서 반영).
            # 아래 값들은 worker._ensure_portfolio_complex /
            # repository.resolve_type_for_import가 소비한다.
            "writer_introduction": _str(row.get("writer_introduction")),
            "real_area_pyeong": row.get("real_area_pyeong"),
            "area_type_structured": _str(row.get("area_type_structured")),
            "area_type_text_detected": _str(row.get("area_type_text_detected")),
            "area_type_confidence": _str(row.get("area_type_confidence")),
            "area_type_needs_lookup": row.get("area_type_needs_lookup"),
            "apartment_name_confidence": _str(row.get("apartment_name_confidence")),
            "street_address_confidence": _str(row.get("street_address_confidence")),
            "address_lookup_confidence": _str(row.get("address_lookup_confidence")),
            "area_id": _str(row.get("area_id")),
            "dong": _str(row.get("dong")),
            "ho": _str(row.get("ho")),
            "expertise": _str(row.get("expertise")),
            "agent": _str(row.get("agent")),
            "family": _str(row.get("family")),
            "style": _str(row.get("style")) or " | ".join(style_tags),
            "construction": _str(row.get("construction")) or " | ".join(construction_tags),
            "period": _int(row.get("period")),
            "budget": _int(row.get("budget")),
            "room_number": _str(row.get("room_number")),
            "created_at": row.get("created_at"),
            "updated_at": row.get("updated_at"),
            "view_count": _int(row.get("view_count")),
            "like_count": _int(row.get("like_count")),
            "spaces": spaces_by_portfolio.get(pid, []),
            "images": images_by_portfolio.get(pid, []),
            "paragraphs": paragraphs_by_portfolio.get(pid, []),
        })

    companies = []
    for row in company_rows:
        companies.append({
            "expert_id": _str(row.get("expert_id")),
            "expert_url": _str(row.get("expert_url")),
            "writer_id": _str(row.get("writer_id")),
            "writer_nickname": _str(row.get("writer_nickname")),
            "company_name": _str(row.get("company_name")),
            "representative_name": _str(row.get("representative_name")),
            "phone": _str(row.get("phone")),
            "email": _str(row.get("email")),
            "address": _str(row.get("address")),
            "business_registration_number": _str(row.get("business_registration_number")),
            "website": _str(row.get("website")),
        })

    return {
        "source": "excel",
        "schema_version": "excel-v1",
        "companies": companies,
        "portfolios": portfolios,
    }

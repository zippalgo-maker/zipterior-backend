import json
import re
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.modules.bulk_import import repository
from app.modules.bulk_import.excel_portfolio import (
    ExcelPortfolioFormatError,
    load_portfolio_workbook,
)
from app.modules.bulk_import.mapping import (
    _excluded_photo_keys,
    build_confidence,
    grouped_portfolio_spaces,
    select_portfolio_images,
    source_space_key,
)
from app.modules.bulk_import.schemas import (
    BulkUploadCreateRequest,
    ComplexResolutionRequest,
)


IMPORT_ROOT = Path("/srv/zipterior/imports")
MAX_CHUNK_SIZE = 10 * 1024 * 1024
MAX_COMPLEX_ROWS = 20_000


class BulkImportNotFoundError(ValueError):
    pass


class BulkImportStateError(ValueError):
    pass


class BulkImportValidationError(ValueError):
    pass


def _job_directory(job_id: int) -> Path:
    return IMPORT_ROOT / str(job_id)


def _normalized_header(value: Any) -> str:
    return re.sub(r"[\s_\-·]+", "", str(value or "")).lower()


def _normalized_key(value: Any) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", str(value or "").lower())


def _extension_for(job_type: str) -> str:
    return ".xlsx" if job_type in ("complex_excel", "company_portfolio_excel") else ".json"


def _require_owned_job(session: Session, *, job_id: int, admin_user_id: int) -> dict[str, Any]:
    job = repository.find_job(session, job_id=job_id, requested_by=admin_user_id)
    if job is None:
        raise BulkImportNotFoundError("일괄등록 작업을 찾을 수 없습니다.")
    return job


class BulkImportService:
    @staticmethod
    def create_upload(
        session: Session,
        *,
        admin_user_id: int,
        payload: BulkUploadCreateRequest,
    ) -> dict[str, Any]:
        expected_extension = _extension_for(payload.job_type)
        if Path(payload.filename).suffix.lower() != expected_extension:
            raise BulkImportValidationError(
                f"{expected_extension} 파일만 업로드할 수 있습니다."
            )
        options = {
            "max_images_per_portfolio": payload.max_images_per_portfolio,
            "max_portfolios": payload.max_portfolios,
            "prefer_complex_address": payload.prefer_complex_address,
            "publish_immediately": payload.publish_immediately,
            "confidence_threshold": payload.confidence_threshold,
            "expert_filter": "agent=전문가",
        }
        job_id = repository.create_job(
            session,
            job_type=payload.job_type,
            filename=payload.filename,
            expected_size=payload.size_bytes,
            options=options,
            requested_by=admin_user_id,
        )
        directory = _job_directory(job_id)
        try:
            directory.mkdir(parents=True, exist_ok=False)
            source_path = directory / f"source{expected_extension}.part"
            source_path.touch(exist_ok=False)
            repository.update_job(
                session,
                job_id=job_id,
                changes={"source_path": str(source_path)},
            )
            session.commit()
        except Exception:
            session.rollback()
            raise
        return repository.find_job(session, job_id=job_id)

    @staticmethod
    async def append_chunk(
        session: Session,
        *,
        admin_user_id: int,
        job_id: int,
        offset: int,
        upload: UploadFile,
    ) -> dict[str, Any]:
        job = repository.lock_job(session, job_id=job_id, requested_by=admin_user_id)
        if job is None:
            raise BulkImportNotFoundError("일괄등록 작업을 찾을 수 없습니다.")
        if job["status"] != "uploading":
            raise BulkImportStateError("업로드 중인 작업만 파일을 이어 올릴 수 있습니다.")
        if offset != int(job["uploaded_size"]):
            raise BulkImportStateError(
                f"업로드 위치가 일치하지 않습니다. 현재 위치는 {job['uploaded_size']}입니다."
            )
        chunk = await upload.read(MAX_CHUNK_SIZE + 1)
        if not chunk:
            raise BulkImportValidationError("비어 있는 조각은 업로드할 수 없습니다.")
        if len(chunk) > MAX_CHUNK_SIZE:
            raise BulkImportValidationError("파일 조각은 최대 10MB까지 허용됩니다.")
        next_size = offset + len(chunk)
        if next_size > int(job["expected_size"]):
            raise BulkImportValidationError("예정된 파일 크기를 초과했습니다.")
        source_path = Path(job["source_path"])
        if not source_path.is_file() or source_path.parent != _job_directory(job_id):
            raise BulkImportStateError("업로드 임시파일을 확인할 수 없습니다.")
        with source_path.open("ab") as stream:
            stream.write(chunk)
        repository.update_job(
            session,
            job_id=job_id,
            changes={"uploaded_size": next_size},
        )
        session.commit()
        return repository.find_job(session, job_id=job_id)

    @staticmethod
    def complete_upload(
        session: Session,
        *,
        admin_user_id: int,
        job_id: int,
    ) -> dict[str, Any]:
        job = _require_owned_job(
            session, job_id=job_id, admin_user_id=admin_user_id
        )
        if job["status"] != "uploading":
            raise BulkImportStateError("업로드 중인 작업만 완료할 수 있습니다.")
        if int(job["uploaded_size"]) != int(job["expected_size"]):
            raise BulkImportStateError("파일 업로드가 아직 완료되지 않았습니다.")
        part_path = Path(job["source_path"])
        final_path = part_path.with_suffix("")
        part_path.replace(final_path)
        try:
            max_images_option = job["options"].get("max_images_per_portfolio")
            max_images = int(max_images_option) if max_images_option is not None else None
            if job["job_type"] == "complex_excel":
                summary = BulkImportService._preview_complex_excel(
                    session, job_id=job_id, source_path=final_path
                )
                status = "awaiting_resolution"
            else:
                preview_fn = (
                    BulkImportService._preview_company_portfolio_excel
                    if job["job_type"] == "company_portfolio_excel"
                    else BulkImportService._preview_company_portfolio_json
                )
                summary = preview_fn(
                    session,
                    job_id=job_id,
                    source_path=final_path,
                    max_images=max_images,
                    max_portfolios=int(job["options"].get("max_portfolios", 30)),
                    prefer_complex_address=bool(
                        job["options"].get("prefer_complex_address", True)
                    ),
                    confidence_threshold=int(
                        job["options"].get("confidence_threshold", 80)
                    ),
                )
                status = (
                    "awaiting_resolution"
                    if summary["pending_resolution_count"]
                    else "preview"
                )
            counts = repository.record_status_counts(session, job_id=job_id)
            repository.update_job(
                session,
                job_id=job_id,
                changes={
                    "source_path": str(final_path),
                    "status": status,
                    "summary": summary,
                    "total_count": summary["eligible_count"],
                    "resolved_count": counts.get("resolved", 0),
                    "skipped_count": summary.get("skipped_count", 0),
                },
            )
            session.commit()
        except Exception as exc:
            session.rollback()
            repository.update_job(
                session,
                job_id=job_id,
                changes={"status": "failed", "error_message": str(exc)[:2000]},
            )
            session.commit()
            if isinstance(exc, BulkImportValidationError):
                raise
            raise BulkImportValidationError(f"파일을 분석하지 못했습니다: {exc}") from exc
        return repository.find_job(session, job_id=job_id)

    @staticmethod
    def _preview_complex_excel(
        session: Session,
        *,
        job_id: int,
        source_path: Path,
    ) -> dict[str, Any]:
        try:
            workbook = load_workbook(source_path, read_only=True, data_only=True)
        except Exception as exc:
            raise BulkImportValidationError("정상적인 Excel(.xlsx) 파일이 아닙니다.") from exc
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise BulkImportValidationError("Excel 첫 행에 컬럼명이 필요합니다.")
        header_map = {_normalized_header(value): index for index, value in enumerate(headers)}
        address_index = next((header_map[key] for key in (
            "주소", "도로명주소", "roadaddress", "address"
        ) if key in header_map), None)
        name_index = next((header_map[key] for key in (
            "건물명", "단지명", "아파트명", "buildingname", "name"
        ) if key in header_map), None)
        if address_index is None:
            raise BulkImportValidationError(
                "주소 또는 도로명 주소 컬럼을 찾을 수 없습니다."
            )
        eligible = 0
        empty = 0
        duplicate_rows = 0
        for row_number, row in enumerate(rows, start=2):
            if row_number > MAX_COMPLEX_ROWS + 1:
                raise BulkImportValidationError(
                    f"단지 일괄등록은 한 파일에 최대 {MAX_COMPLEX_ROWS:,}행까지 가능합니다."
                )
            address = str(row[address_index] or "").strip() if address_index < len(row) else ""
            name = (
                str(row[name_index] or "").strip()
                if name_index is not None and name_index < len(row)
                else ""
            )
            if not address:
                empty += 1
                continue
            record_key = (_normalized_key(address) + "|" + _normalized_key(name))[:300]
            record_id = repository.create_record(
                session,
                job_id=job_id,
                record_type="complex",
                record_key=record_key,
                source_label=name or address,
                payload={"row_number": row_number, "address": address, "name": name},
            )
            if record_id is None:
                duplicate_rows += 1
            else:
                eligible += 1
        workbook.close()
        return {
            "eligible_count": eligible,
            "empty_address_count": empty,
            "duplicate_row_count": duplicate_rows,
            "skipped_count": empty + duplicate_rows,
            "required_columns": ["주소"],
            "optional_columns": ["건물명"],
        }

    @staticmethod
    def _preview_company_portfolio_json(
        session: Session,
        *,
        job_id: int,
        source_path: Path,
        max_images: int | None,
        max_portfolios: int,
        prefer_complex_address: bool,
        confidence_threshold: int,
    ) -> dict[str, Any]:
        try:
            with source_path.open("r", encoding="utf-8-sig") as stream:
                data = json.load(stream)
        except Exception as exc:
            raise BulkImportValidationError("정상적인 UTF-8 JSON 파일이 아닙니다.") from exc
        if not isinstance(data, dict) or not isinstance(data.get("portfolios"), list):
            raise BulkImportValidationError("portfolios 배열이 있는 JSON 파일이 필요합니다.")
        return BulkImportService._preview_company_portfolio_data(
            session,
            job_id=job_id,
            data=data,
            max_images=max_images,
            max_portfolios=max_portfolios,
            prefer_complex_address=prefer_complex_address,
            confidence_threshold=confidence_threshold,
        )

    @staticmethod
    def _preview_company_portfolio_excel(
        session: Session,
        *,
        job_id: int,
        source_path: Path,
        max_images: int | None,
        max_portfolios: int,
        prefer_complex_address: bool,
        confidence_threshold: int,
    ) -> dict[str, Any]:
        try:
            data = load_portfolio_workbook(source_path)
        except ExcelPortfolioFormatError as exc:
            raise BulkImportValidationError(str(exc)) from exc
        return BulkImportService._preview_company_portfolio_data(
            session,
            job_id=job_id,
            data=data,
            max_images=max_images,
            max_portfolios=max_portfolios,
            prefer_complex_address=prefer_complex_address,
            confidence_threshold=confidence_threshold,
        )

    @staticmethod
    def _preview_company_portfolio_data(
        session: Session,
        *,
        job_id: int,
        data: dict[str, Any],
        max_images: int | None,
        max_portfolios: int,
        prefer_complex_address: bool,
        confidence_threshold: int,
    ) -> dict[str, Any]:
        companies = data.get("companies") if isinstance(data.get("companies"), list) else []
        portfolios = data["portfolios"]
        experts = [item for item in portfolios if str(item.get("agent") or "").strip() == "전문가"]
        if prefer_complex_address:
            address_ready = [
                item for item in experts
                if str(item.get("street_address") or "").strip()
            ]
            address_missing = [
                item for item in experts
                if not str(item.get("street_address") or "").strip()
            ]
            ordered = address_ready + address_missing
        else:
            ordered = experts
        selected = ordered[:max_portfolios]
        images_total = sum(len(item.get("images") or []) for item in selected)
        selected_by_portfolio = [
            select_portfolio_images(item, max_images=max_images)
            for item in selected
        ]
        selected_images = sum(len(images) for images in selected_by_portfolio)
        before_images = sum(
            str(image.get("phase") or "").strip().upper() == "BEFORE"
            for item in selected
            for image in item.get("images") or []
        )
        after_images = images_total - before_images
        unmatched_image_groups = 0
        missed_after_groups = 0
        for item, chosen in zip(selected, selected_by_portfolio, strict=True):
            known_groups = {space["key"] for space in grouped_portfolio_spaces(item)}
            after_groups = {
                source_space_key(image)
                for image in item.get("images") or []
                if str(image.get("phase") or "").strip().upper() == "AFTER"
            }
            chosen_groups = {source_space_key(image) for image in chosen}
            unmatched_image_groups += len(after_groups - known_groups)
            missed_after_groups += len(after_groups - chosen_groups)
        unique_writers = {
            str(item.get("writer_id") or item.get("expert_id") or "").strip()
            for item in selected
            if item.get("writer_id") or item.get("expert_id")
        }
        pending_resolution = 0
        confidence_scored_count = 0
        needs_review_count = 0
        for item in selected:
            source_key = str(item.get("portfolio_id") or "").strip()
            if not source_key:
                continue
            address = str(item.get("street_address") or "").strip()
            status = "pending" if address else "resolved"
            pending_resolution += int(status == "pending")
            # v2.5.0: 구조 신호 기반 신뢰도를 미리보기 단계에서 미리 계산해 두면
            # 관리자가 실제 등록 전에 검수할 수 있다. paragraphs(엑셀 문단 단위
            # 원본)가 없는 소스(구버전 JSON)는 build_confidence가 None을 반환하며,
            # 이 경우 채점 대상이 아니므로 기본 선택(공개)으로 취급한다.
            confidence = build_confidence(item)
            confidence_score = confidence.portfolio_score if confidence else None
            if confidence is not None:
                confidence_scored_count += 1
                if confidence_score < confidence_threshold:
                    needs_review_count += 1
            admin_selected = (
                True if confidence_score is None else confidence_score >= confidence_threshold
            )
            # v2.5.0: 실제 등록(select_portfolio_images/content_blocks_from_item)
            # 에서 제외되는 사진(ETC, 마지막 두 장 text_context)은 등록 전
            # 미리보기 화면에도 똑같이 안 보여야 한다 -- 안 그러면 미리보기와
            # 실제 등록 결과가 서로 다른 사진을 보여주는 모순이 생긴다.
            excluded_card_ids, excluded_urls = _excluded_photo_keys(item)
            repository.create_record(
                session,
                job_id=job_id,
                record_type="portfolio",
                record_key=source_key,
                source_label=str(item.get("title") or source_key)[:500],
                payload={
                    "portfolio_id": source_key,
                    "writer_id": item.get("writer_id"),
                    "title": item.get("title"),
                    "name": str(item.get("apartment_name") or "").strip(),
                    "address": address,
                    "image_count": len(item.get("images") or []),
                    "source_url": item.get("source_url"),
                    "confidence_score": confidence_score,
                    "confidence_available": confidence is not None,
                    "admin_selected": admin_selected,
                    "confidence_sections": (
                        [
                            {
                                "label": s.label,
                                "score": s.score,
                                "reason": s.reason,
                                "n_images": s.n_images,
                                "n_text_paragraphs": s.n_text_paragraphs,
                                "text": (s.text or "")[:2000],
                            }
                            for s in confidence.sections
                        ]
                        if confidence
                        else []
                    ),
                    "intro_text": (confidence.intro_text or "")[:2000] if confidence else "",
                    "closing_text": (confidence.closing_text or "")[:2000] if confidence else "",
                    "used_structural_headings": (
                        confidence.used_structural_headings if confidence else None
                    ),
                    "platform_mentions_removed": (
                        confidence.platform_mentions_removed if confidence else 0
                    ),
                    # v2.5.0: 검수 화면에서 방/사진별 실제 미리보기를 렌더링할 수 있게
                    # 사진 URL과 이름표를 그대로 보존한다(다운로드 전, 원본 링크).
                    "preview_images": [
                        {
                            "url": img.get("image_url"),
                            "label": str(img.get("sub_space_name") or img.get("space_name") or "").strip(),
                            "order": img.get("document_order") or img.get("image_order"),
                        }
                        for img in (item.get("images") or [])
                        if str(img.get("phase") or "").strip().upper() != "BEFORE"
                        and img.get("image_url")
                        and img.get("ext_card_id") not in excluded_card_ids
                        and str(img.get("image_url") or "").strip() not in excluded_urls
                    ][:200],
                },
                status=status,
            )
        return {
            "source": str(data.get("source") or "unknown")[:100],
            "schema_version": str(data.get("schema_version") or "")[:50],
            "company_count": len(companies),
            "portfolio_count": len(portfolios),
            "eligible_count": len(selected),
            "expert_total_count": len(experts),
            "skipped_count": len(portfolios) - len(selected),
            "non_expert_excluded_count": len(portfolios) - len(experts),
            "sample_limit_excluded_count": max(0, len(experts) - len(selected)),
            "unique_expert_writer_count": len(unique_writers),
            "image_url_count": images_total,
            "after_image_count": after_images,
            "selected_image_count": selected_images,
            "image_limit_excluded_count": after_images - selected_images,
            "before_image_excluded_count": before_images,
            "unmatched_image_group_count": unmatched_image_groups,
            "missed_after_group_count": missed_after_groups,
            "pending_resolution_count": pending_resolution,
            "selected_with_address_count": sum(
                bool(str(item.get("street_address") or "").strip())
                for item in selected
            ),
            "max_portfolios": max_portfolios,
            "prefer_complex_address": prefer_complex_address,
            "expert_filter": "agent=전문가",
            "confidence_threshold": confidence_threshold,
            "confidence_scored_count": confidence_scored_count,
            "confidence_needs_review_count": needs_review_count,
        }

    @staticmethod
    def resolve_complexes(
        session: Session,
        *,
        admin_user_id: int,
        job_id: int,
        payload: ComplexResolutionRequest,
    ) -> dict[str, Any]:
        if not payload.items and not payload.failures:
            raise BulkImportValidationError("주소 확인 결과가 필요합니다.")
        job = _require_owned_job(
            session, job_id=job_id, admin_user_id=admin_user_id
        )
        if job["status"] != "awaiting_resolution":
            raise BulkImportStateError("주소 확인 중인 일괄등록 작업이 아닙니다.")
        expected_record_type = (
            "complex" if job["job_type"] == "complex_excel" else "portfolio"
        )
        for item in payload.items:
            record = repository.find_record(session, job_id=job_id, record_id=item.record_id)
            if record is None or record["record_type"] != expected_record_type:
                raise BulkImportNotFoundError("주소 확인 대상을 찾을 수 없습니다.")
            merged = {**record["payload"], **item.model_dump(exclude={"record_id"})}
            repository.update_record(
                session,
                record_id=item.record_id,
                changes={
                    "status": "resolved",
                    "payload": merged,
                    "source_label": (
                        item.name
                        if expected_record_type == "complex"
                        else record["source_label"]
                    ),
                    "error_message": None,
                },
            )
        for item in payload.failures:
            record = repository.find_record(session, job_id=job_id, record_id=item.record_id)
            if record is None or record["record_type"] != expected_record_type:
                raise BulkImportNotFoundError("주소 확인 대상을 찾을 수 없습니다.")
            if expected_record_type == "portfolio":
                merged = {
                    **record["payload"],
                    "address_resolution_error": item.error_message,
                }
                changes = {
                    "status": "resolved",
                    "payload": merged,
                    "error_message": item.error_message,
                }
            else:
                changes = {"status": "failed", "error_message": item.error_message}
            repository.update_record(
                session,
                record_id=item.record_id,
                changes=changes,
            )
        counts = repository.record_status_counts(session, job_id=job_id)
        next_status = job["status"]
        if job["job_type"] in ("company_portfolio_json", "company_portfolio_excel") and not counts.get("pending", 0):
            next_status = "preview"
        repository.update_job(
            session,
            job_id=job_id,
            changes={
                "resolved_count": counts.get("resolved", 0),
                "failed_count": counts.get("failed", 0),
                "status": next_status,
            },
        )
        session.commit()
        return repository.find_job(session, job_id=job_id)

    @staticmethod
    def start_job(
        session: Session,
        *,
        admin_user_id: int,
        job_id: int,
    ) -> dict[str, Any]:
        job = _require_owned_job(
            session, job_id=job_id, admin_user_id=admin_user_id
        )
        allowed = {"preview"}
        if job["job_type"] == "complex_excel":
            allowed.add("awaiting_resolution")
            counts = repository.record_status_counts(session, job_id=job_id)
            if counts.get("pending", 0):
                raise BulkImportStateError("모든 주소의 카카오 확인을 먼저 완료해 주세요.")
            if not counts.get("resolved", 0):
                raise BulkImportStateError("등록 가능한 단지 주소가 없습니다.")
        if job["status"] not in allowed:
            raise BulkImportStateError("미리보기 완료 상태에서만 일괄등록을 시작할 수 있습니다.")
        repository.update_job(
            session,
            job_id=job_id,
            changes={"status": "queued", "error_message": None},
        )
        session.commit()
        return repository.find_job(session, job_id=job_id)

    @staticmethod
    def cancel_job(
        session: Session,
        *,
        admin_user_id: int,
        job_id: int,
    ) -> dict[str, Any]:
        job = _require_owned_job(
            session, job_id=job_id, admin_user_id=admin_user_id
        )
        if job["status"] in {"completed", "completed_with_errors", "failed", "cancelled"}:
            raise BulkImportStateError("이미 종료된 작업입니다.")
        repository.update_job(
            session,
            job_id=job_id,
            changes={"status": "cancelled"},
        )
        # completed_at은 SQL 함수 문자열이 아니라 DB 기본 시각으로 별도 갱신한다.
        session.execute(
            text("UPDATE bulk_import_jobs SET completed_at=NOW() WHERE id=:job_id"),
            {"job_id": job_id},
        )
        session.commit()
        return repository.find_job(session, job_id=job_id)

    @staticmethod
    def get_job(session: Session, *, admin_user_id: int, job_id: int) -> dict[str, Any]:
        return _require_owned_job(session, job_id=job_id, admin_user_id=admin_user_id)

    @staticmethod
    def list_jobs(session: Session, *, admin_user_id: int, limit: int) -> list[dict[str, Any]]:
        return repository.list_jobs(session, requested_by=admin_user_id, limit=limit)

    @staticmethod
    def list_records(
        session: Session,
        *,
        admin_user_id: int,
        job_id: int,
        status: str | None,
        limit: int,
        offset: int,
    ) -> tuple[list[dict[str, Any]], int]:
        # 2026-08-25: 상세 화면 결과 목록이 total 없이 limit=500 한
        # 페이지로만 통째로 내려가서, 500건 넘는 대량 작업(실제로 458건
        # 까지 있었음)은 그 이상이 조용히 잘리고 있었다(사용자 리포트:
        # "페이지 넘기는 게 없다"). items와 total을 같이 반환해 라우터가
        # 진짜 페이지네이션 응답을 내려주게 함 -- worker.py 등 내부
        # 처리 루프는 이 서비스 메서드가 아니라 repository.list_records를
        # 직접 호출하므로 영향 없음.
        _require_owned_job(session, job_id=job_id, admin_user_id=admin_user_id)
        items = repository.list_records(
            session, job_id=job_id, status=status, limit=limit, offset=offset
        )
        total = repository.count_records(session, job_id=job_id, status=status)
        return items, total

    @staticmethod
    def set_record_selection(
        session: Session,
        *,
        admin_user_id: int,
        job_id: int,
        record_id: int,
        selected: bool,
    ) -> dict[str, Any]:
        """v2.5.0 관리자 검수: 신뢰도 자동판정과 별개로 관리자가 개별 레코드의
        공개 여부 체크를 자유롭게 뒤집을 수 있게 한다. 미리보기/주소확인
        단계에서만 허용 -- 이미 시작된 작업의 결과를 건드리지 않는다."""
        job = _require_owned_job(session, job_id=job_id, admin_user_id=admin_user_id)
        if job["status"] not in {"preview", "awaiting_resolution"}:
            raise BulkImportStateError("미리보기 단계에서만 선택 상태를 변경할 수 있습니다.")
        record = repository.find_record(session, job_id=job_id, record_id=record_id)
        if record is None:
            raise BulkImportNotFoundError("레코드를 찾을 수 없습니다.")
        if record["record_type"] != "portfolio":
            raise BulkImportValidationError("포트폴리오 레코드만 선택 상태를 변경할 수 있습니다.")
        payload = dict(record["payload"])
        payload["admin_selected"] = bool(selected)
        repository.update_record(session, record_id=record_id, changes={"payload": payload})
        session.commit()
        return repository.find_record(session, job_id=job_id, record_id=record_id)

    @staticmethod
    def apply_confidence_threshold(
        session: Session,
        *,
        admin_user_id: int,
        job_id: int,
        threshold: int,
    ) -> dict[str, Any]:
        """v2.5.0 관리자 검수: 기준값(%)을 바꿔서 모든 포트폴리오 레코드의
        기본 선택 상태를 일괄 재계산한다. 개별로 뒤집어 둔 체크도 이 시점에는
        기준값 기준으로 초기화된다 -- 재계산 이후 다시 개별 조정하면 된다."""
        job = _require_owned_job(session, job_id=job_id, admin_user_id=admin_user_id)
        if job["status"] not in {"preview", "awaiting_resolution"}:
            raise BulkImportStateError("미리보기 단계에서만 기준값을 조정할 수 있습니다.")
        records = repository.list_records(
            session, job_id=job_id, status=None, limit=10_000, offset=0
        )
        for record in records:
            if record["record_type"] != "portfolio":
                continue
            payload = dict(record["payload"])
            score = payload.get("confidence_score")
            payload["admin_selected"] = True if score is None else score >= threshold
            repository.update_record(session, record_id=record["id"], changes={"payload": payload})
        options = dict(job["options"])
        options["confidence_threshold"] = int(threshold)
        summary = dict(job["summary"])
        summary["confidence_threshold"] = int(threshold)
        summary["confidence_needs_review_count"] = sum(
            1
            for record in records
            if record["record_type"] == "portfolio"
            and record["payload"].get("confidence_score") is not None
            and record["payload"].get("confidence_score") < threshold
        )
        repository.update_job(
            session, job_id=job_id, changes={"options": options, "summary": summary}
        )
        session.commit()
        return repository.find_job(session, job_id=job_id)
